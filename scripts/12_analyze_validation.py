"""
Analyze the filled validation workbooks (run AFTER both annotators finish).

Reads the two annotator .xlsx files from the Drive validation folder and the
blinding key (data/validation_sample_key.json), then reports:
  - Cohen's kappa between annotators (Q1, 3 categories and binarized)
  - Consensus human label (agreements; disagreements -> adjudicate or drop)
  - Heuristic vs human: precision, recall, F1, accuracy (binary breakdown)
  - Among true breakdowns: repaired vs persisted (responds to R1.2)
  - Breakdown of errors by flag kind (marker vs chain) and tier group

Usage: python3 scripts/12_analyze_validation.py
"""

import json
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DRIVE = Path("/mnt/g/My Drive/COPPEAD/papers/the-communication-ceiling/v2/validation")
KEY = json.loads((ROOT / "data" / "validation_sample_key.json").read_text())
# active subset (author decision 2026-08-12: n=120; the other 80 of the
# original 200 are a blind reserve — see scripts/14_subset_validation_120.py)
SUBSET_FILE = ROOT / "data" / "validation_subset_120.json"


def normalize(v):
    if not v:
        return None
    v = str(v).strip().lower()
    # tolerate accented/variant spellings from manual entry
    v = (v.replace("ã", "a").replace("á", "a").replace("â", "a")
          .replace("é", "e").replace("ê", "e").replace("í", "i")
          .replace("ó", "o").replace("ô", "o").replace("ú", "u")
          .replace("ç", "c"))
    return v or None


def load_annotations(path):
    from openpyxl import load_workbook
    ws = load_workbook(path, read_only=True)["Conversas"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None] * 6
        seq, q1, q2 = row[0], row[3], row[4]
        if seq is None:
            continue
        out[int(seq)] = (normalize(q1), normalize(q2))
    return out


def kappa(pairs):
    n = len(pairs)
    if n == 0:
        return None
    po = sum(1 for a, b in pairs if a == b) / n
    ca, cb = Counter(a for a, _ in pairs), Counter(b for _, b in pairs)
    pe = sum((ca[k] / n) * (cb[k] / n) for k in set(ca) | set(cb))
    return (po - pe) / (1 - pe) if pe < 1 else None


def main():
    A = load_annotations(DRIVE / "validation-ANOTADOR-A_Preenchido Alex.xlsx")
    B = load_annotations(DRIVE / "validation-ANOTADOR-B_Preenchido Eduardo.xlsx")
    items = {it["seq"]: it for it in KEY["items"]}
    if SUBSET_FILE.exists():
        subset = set(json.loads(SUBSET_FILE.read_text())["seqs"])
        items = {s: it for s, it in items.items() if s in subset}
        print(f"active subset: n={len(items)} (blind reserve: "
              f"{len(KEY['items']) - len(items)})")

    answered = [s for s in items if A.get(s, (None,))[0] and B.get(s, (None,))[0]]
    # conversations either annotator could not judge (e.g. mislabeled language)
    # are excluded from kappa/consensus but reported: they estimate the rate of
    # non-judgeable items that the corpus language filter lets through.
    not_rateable = [s for s in answered
                    if "nao-avaliavel" in (A[s][0], B[s][0])]
    both = [s for s in answered if s not in not_rateable]
    print(f"answered by both: {len(answered)}/{len(items)}")
    print(f"nao-avaliavel for at least one annotator: {len(not_rateable)} "
          f"-> seqs {not_rateable} (excluded from kappa/consensus)")
    if not both:
        print("Nothing to analyze yet."); return

    pairs3 = [(A[s][0], B[s][0]) for s in both]
    to_bin = lambda v: "sim" if v == "sim" else "nao"  # incerto -> nao (conservative)
    pairs2 = [(to_bin(a), to_bin(b)) for a, b in pairs3]
    print(f"kappa (3 categories): {kappa(pairs3):.3f}")
    print(f"kappa (binary, incerto->nao): {kappa(pairs2):.3f}")

    # consensus: both agree on binary label; disagreements listed for adjudication
    disagreements = [s for s in both if to_bin(A[s][0]) != to_bin(B[s][0])]
    print(f"binary disagreements to adjudicate: {len(disagreements)} -> seqs {disagreements[:30]}")

    consensus = {s: to_bin(A[s][0]) for s in both if s not in disagreements}
    tp = fp = fn = tn = 0
    err_by_kind = Counter()
    for s, human in consensus.items():
        flagged = items[s]["flag_kind"] in ("marker", "chain")
        if flagged and human == "sim":
            tp += 1
        elif flagged and human == "nao":
            fp += 1; err_by_kind[("fp", items[s]["flag_kind"])] += 1
        elif not flagged and human == "sim":
            fn += 1; err_by_kind[("fn", "clean")] += 1
        else:
            tn += 1
    prec = tp / (tp + fp) if tp + fp else float("nan")
    rec = tp / (tp + fn) if tp + fn else float("nan")
    f1 = 2 * prec * rec / (prec + rec) if prec + rec else float("nan")
    print(f"\nheuristic vs human consensus (n={len(consensus)}):")
    print(f"  TP={tp} FP={fp} FN={fn} TN={tn}")
    print(f"  precision={prec:.3f} recall={rec:.3f} F1={f1:.3f} "
          f"accuracy={(tp+tn)/len(consensus):.3f}")
    print(f"  errors by kind: {dict(err_by_kind)}")
    print("  NOTE: sample is 50/50 stratified, NOT the corpus base rate; report "
          "precision/recall per stratum, not raw accuracy, in the paper.")

    # R1.2: among human-confirmed breakdowns, repaired vs persisted (Q2)
    q2 = Counter()
    for s in consensus:
        if consensus[s] == "sim":
            for ann in (A, B):
                v = ann[s][1]
                if v:
                    q2[v] += 1
    print(f"\nQ2 among confirmed breakdowns (both annotators pooled): {dict(q2)}")


if __name__ == "__main__":
    main()
