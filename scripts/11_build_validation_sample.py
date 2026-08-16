"""
Build the human-annotation validation sample (Frontiers R1.1-R1.2, R2.1-R2.2).

Design (mirrors the Session-1 validation plan, stratified):
  - 200 WildChat conversations from the 6-10 turn stratum (the paper's
    reporting stratum), in four groups of 50:
      G1 flagged   x GPT-3.5        G2 flagged   x GPT-4-or-later
      G3 unflagged x GPT-3.5        G4 unflagged x GPT-4-or-later
  - Within each flagged group, up to 20 conversations flagged by explicit
    failure MARKERS (rarer) and the rest by refinement CHAINS only.
  - Reservoir sampling with a fixed seed (42) for reproducibility.

Outputs:
  - data/validation_sample_key.json  (IDs + true flags — KEEP AWAY from annotators)
  - two identical blinded Excel workbooks (instructions + conversations +
    dropdown answer columns), written to the Drive validation folder.

Annotators judge, per conversation:
  Q1  breakdown  : did mutual understanding observably fail at any point?
  Q2  resolution : if yes, was it repaired by the end? (repaired/persisted/unclear)
"""

import json
import random
import time
from pathlib import Path
import importlib.util

ROOT = Path(__file__).resolve().parent.parent
spec = importlib.util.spec_from_file_location(
    "s03", ROOT / "scripts" / "03_failure_by_model.py")
s03 = importlib.util.module_from_spec(spec)
spec.loader.exec_module(s03)

SEED = 42
PER_GROUP = 50
MARKER_TARGET = 20  # per flagged group
# Per-message caps keep the FULL turn sequence visible within Excel's
# 32,767-char cell limit (worst case ~20 messages x ~1,230 chars ~= 26k).
ASSIST_TRUNC = 1200
USER_TRUNC = 1200

DRIVE_DIR = Path("/mnt/g/My Drive/COPPEAD/papers/the-communication-ceiling/v2/validation")
KEY_FILE = ROOT / "data" / "validation_sample_key.json"

GPT35 = {m for m, t in s03.MODEL_TIERS.items() if t == "gpt-3.5"}
GPT4P = {m for m, t in s03.MODEL_TIERS.items()
         if t in ("gpt-4", "gpt-4o", "gpt-4o-mini", "gpt-4.1-mini")}


def render(conv):
    parts = []
    for msg in conv:
        role = msg.get("role", "")
        content = msg.get("content", "")
        if not isinstance(content, str):
            continue
        if role == "assistant" and len(content) > ASSIST_TRUNC:
            content = content[:ASSIST_TRUNC] + " [... resposta truncada ...]"
        elif role != "assistant" and len(content) > USER_TRUNC:
            content = content[:USER_TRUNC] + " [... mensagem truncada ...]"
        parts.append(("USER" if role == "user" else "ASSISTANT") + ": " + content.strip())
    return "\n\n".join(parts)


class Reservoir:
    def __init__(self, k, rng):
        self.k, self.rng, self.items, self.seen = k, rng, [], 0

    def offer(self, item):
        self.seen += 1
        if len(self.items) < self.k:
            self.items.append(item)
        else:
            j = self.rng.randrange(self.seen)
            if j < self.k:
                self.items[j] = item


def main():
    import pandas as pd

    rng = random.Random(SEED)
    pools = {  # (tier_group, flag_kind) -> reservoir
        ("gpt35", "marker"): Reservoir(MARKER_TARGET, rng),
        ("gpt35", "chain"): Reservoir(PER_GROUP, rng),
        ("gpt35", "clean"): Reservoir(PER_GROUP, rng),
        ("gpt4p", "marker"): Reservoir(MARKER_TARGET, rng),
        ("gpt4p", "chain"): Reservoir(PER_GROUP, rng),
        ("gpt4p", "clean"): Reservoir(PER_GROUP, rng),
    }

    t0 = time.time()
    scanned = 0
    for file_idx in range(86):
        pf = s03.CACHE_DIR / f"train-{file_idx:05d}-of-00086.parquet"
        if not pf.exists():
            continue
        df = pd.read_parquet(pf, columns=[
            "conversation_hash", "model", "conversation", "language"])
        eng = df[df["language"] == "English"]
        for row in eng.itertuples(index=False):
            model = row.model or ""
            if model in GPT35:
                tg = "gpt35"
            elif model in GPT4P:
                tg = "gpt4p"
            else:
                continue
            conv = row.conversation
            if conv is None:
                continue
            if not isinstance(conv, list):
                try:
                    conv = list(conv)
                except Exception:
                    continue
            res = s03.analyze_conversation(conv)
            if res is None or not (6 <= res["n_user_turns"] <= 10):
                continue
            scanned += 1
            if res["has_failure_markers"]:
                kind = "marker"
            elif res["has_refinement_chains"]:
                kind = "chain"
            else:
                kind = "clean"
            pools[(tg, kind)].offer({
                "id": row.conversation_hash, "model": model, "tier_group": tg,
                "flag_kind": kind, "n_user_turns": res["n_user_turns"],
                "text": render(conv),
            })
        if file_idx % 10 == 9:
            print(f"  [{file_idx+1}/86] eligible={scanned:,} {time.time()-t0:.0f}s",
                  flush=True)

    sample = []
    for tg in ("gpt35", "gpt4p"):
        markers = pools[(tg, "marker")].items
        chains = pools[(tg, "chain")].items
        flagged = markers + chains[: PER_GROUP - len(markers)]
        clean = pools[(tg, "clean")].items[:PER_GROUP]
        sample.extend(flagged)
        sample.extend(clean)
        print(f"{tg}: flagged={len(flagged)} (markers={len(markers)}) clean={len(clean)}")

    rng.shuffle(sample)
    for i, item in enumerate(sample, 1):
        item["seq"] = i

    KEY_FILE.write_text(json.dumps({
        "metadata": {"seed": SEED, "stratum": "6-10 turns", "n": len(sample),
                     "eligible_scanned": scanned,
                     "design": "50 flagged + 50 clean per tier group (GPT-3.5 / GPT-4+)"},
        "items": [{k: v for k, v in it.items() if k != "text"} for it in sample],
    }, indent=2))

    # ---- Excel workbooks -------------------------------------------------
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    INSTRUCTIONS = [
        ("Validação humana — Falhas de comunicação (communication breakdown)", True),
        ("", False),
        ("O QUE FAZER", True),
        ("Leia cada conversa (aba 'Conversas') entre um usuário e um assistente de IA e responda às duas perguntas nas colunas D e E.", False),
        ("", False),
        ("Q1 — BREAKDOWN (coluna D): Em algum ponto da conversa a compreensão mútua falhou de forma observável? Ou seja: o assistente entendeu errado o pedido, o usuário precisou corrigir, reformular ou expressar que foi mal-entendido? Responda: sim / nao / incerto / nao-avaliavel.", False),
        ("Use 'nao-avaliavel' quando não for possível julgar a conversa — por exemplo, quando ela estiver em um idioma que você não lê. Nesse caso, deixe a Q2 em branco e passe para a próxima.", False),
        ("", False),
        ("Q2 — RESOLUCAO (coluna E): Responda APENAS se Q1 = sim. Ao final da conversa, o mal-entendido foi consertado? reparado = o usuário conseguiu o que queria após a correção; persistiu = a conversa termina sem resolver; incerto = impossível dizer.", False),
        ("", False),
        ("NOTAS (coluna F): opcional, qualquer observação.", False),
        ("", False),
        ("REGRAS IMPORTANTES", True),
        ("1. NÃO discuta as conversas com a outra pessoa anotadora até que ambas terminem.", False),
        ("2. Julgue APENAS falha de COMUNICAÇÃO (entendimento), não a qualidade da resposta. Uma resposta correta porém mal escrita NÃO é breakdown; uma resposta que revela que o pedido foi mal entendido É.", False),
        ("3. Reformulação para MELHORAR estilo (ex.: 'na verdade, encurte isso') sem mal-entendido NÃO é breakdown.", False),
        ("4. Conversas sobre programação: julgue a COMUNICAÇÃO (o assistente entendeu o que foi pedido?), não a corretude técnica do código. Código com erro não é breakdown por si só; o usuário precisar re-explicar o que pediu É.", False),
        ("5. Mensagens muito longas foram encurtadas ('[... mensagem truncada ...]' / '[... resposta truncada ...]'), mas TODOS os turnos da conversa estão visíveis; julgue pelo que está visível.", False),
        ("6. As conversas vêm de um corpus público real (em inglês) e podem conter conteúdo estranho ou sensível; anote normalmente.", False),
        ("", False),
        ("Tempo estimado: 2-4 minutos por conversa.", False),
    ]

    def build_workbook(path, annotator_label):
        wb = Workbook()
        ws = wb.active
        ws.title = "Instrucoes"
        ws.column_dimensions["A"].width = 120
        for i, (text, bold) in enumerate(INSTRUCTIONS, 1):
            c = ws.cell(row=i, column=1, value=text)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if bold:
                c.font = Font(bold=True)

        ws2 = wb.create_sheet("Conversas")
        headers = ["#", "ID", "Conversa (USER/ASSISTANT)",
                   "Q1 breakdown (sim/nao/incerto/nao-avaliavel)",
                   "Q2 resolucao (reparado/persistiu/incerto)",
                   "Notas"]
        widths = [6, 14, 110, 24, 26, 30]
        fill = PatternFill("solid", fgColor="1D3557")
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="center")
            ws2.column_dimensions[chr(64 + col)].width = w
        ws2.freeze_panes = "A2"

        dv1 = DataValidation(type="list", formula1='"sim,nao,incerto,nao-avaliavel"',
                             allow_blank=True, showDropDown=False)
        dv2 = DataValidation(type="list", formula1='"reparado,persistiu,incerto"',
                             allow_blank=True, showDropDown=False)
        ws2.add_data_validation(dv1)
        ws2.add_data_validation(dv2)

        for r, item in enumerate(sample, 2):
            ws2.cell(row=r, column=1, value=item["seq"])
            ws2.cell(row=r, column=2, value=str(item["id"])[:12])
            c = ws2.cell(row=r, column=3, value=item["text"][:32000])
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws2.row_dimensions[r].height = 220
            dv1.add(ws2.cell(row=r, column=4))
            dv2.add(ws2.cell(row=r, column=5))

        ws.cell(row=len(INSTRUCTIONS) + 2, column=1,
                value=f"Planilha: {annotator_label}")
        wb.save(path)
        print("saved", path)

    DRIVE_DIR.mkdir(parents=True, exist_ok=True)
    build_workbook(DRIVE_DIR / "validation-ANOTADOR-A.xlsx", "Anotador A")
    build_workbook(DRIVE_DIR / "validation-ANOTADOR-B.xlsx", "Anotador B")
    print(f"\nSample n={len(sample)} | key: {KEY_FILE}")


if __name__ == "__main__":
    main()
